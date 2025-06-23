from django.shortcuts import render, redirect
from datetime import datetime
from home.models import Post
from django.http import JsonResponse
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from django.http import HttpResponse
from io import BytesIO
from PIL import Image
import time
import pytesseract
import re
import os

# Create your views here.
def homepage(request):
    posts = Post.objects.all()
    now = datetime.now()
    return render(request, "home.html", locals())


def showpost(request, slug):
    try:
        post = Post.objects.get(slug=slug)
        if post != None:
            return render(request, "post.html", locals())
    except:
        return redirect("/")

def ppsg(requset):
     
    account = '帳號：'
    pswd = '密碼：'
    game = '遊戲類別：'
    environment = '測試環境：'
    excel = '上傳檔案：'
    website = '網站：'
    
    context = {
        'account': account,
        'pswd': pswd,
        'game': game,
        'environment': environment,
        'excel': excel,
        'website':website
    }
    
    return render(requset, "ppsg.html", context)

def upload(request):
    global game
    global gamename
    global environment
    global uploaded_file
    
    if request.method == 'POST':
        account = request.POST.get('account')
        pswd = request.POST.get('pswd')
        game = request.POST.get('game')
        environment = request.POST.get('environment')
        website = request.POST.get('website')
        uploaded_file = request.FILES.getlist('excelFile')  # 獲取上傳的檔案         
        
        gamename = excel_file_name(uploaded_file)
        
        parameters(website, environment, account, pswd)

        return JsonResponse({'message': '檔案上傳成功！'}, status=200)
    else:
        return JsonResponse({'error': '不支持的請求方法'}, status=400)

def excel_file_name(file):
    gamename = []
    for uploaded_file_name in file:
            excel_game_name = uploaded_file_name.name
            excel_game_name = excel_game_name.replace('.xlsx','')
            gamename.append(excel_game_name)
    return gamename

def ppsgFunctionSelection(request):
    posts = Post.objects.all()
    now = datetime.now()
    return render(request, "ppsgFunctionSelection.html", locals())


def excel_list(file_path):
    gameName_list = []
    # excel = load_workbook(file_path)  # 讀取excel檔
    for x in file_path:
        excel = load_workbook(x)
        for sheetName in excel.sheetnames:  # 把excel 的sheet全部取出遍歷
            try:
                sheet = excel[sheetName]  # 對EXCEL切換sheet
            except:
                break
            excelGame = sheet[1]
            excelGameName = excelGame[0].value
            if excelGameName == None:
                continue
            else:
                gameNameRow = sheet[1]
                gameNameValue = gameNameRow[0].value
                gameName_list.append(gameNameValue)
    return gameName_list

def excel_gcadmin(file_path):
    gameTypeGameName = {}
    # excel = load_workbook(file_path)
    excel = load_workbook(file_path)
    sheet = excel["Sheet1"]
    for row in range(2, 100):
        gametype_row = sheet[row]
        gametype_value = gametype_row[0].value
        if gametype_value != None:
            ProviderGameType = gametype_row[1].value
            gameNameEnglishValue = gametype_row[2].value
            gameTypeGameName[gametype_value] = (
                gameNameEnglishValue,
                ProviderGameType,
            )

    return gameTypeGameName

#參數判斷和URL設定
def parameters(website, environment, account, pswd):
    
    url_mapping = {
        ('thor_admin'):'https://admin.12vin.com/(S(h2uux4srsyv0fwgu3ym2pmti))/default.aspx',
        ('thor_agent'):'https://agent.12vin.com/(S(1clf4jfquob2frkapsd1egc3))/default.aspx',
        ('thor_max222agent'):'https://max222agent.12vin.com/(S(kqtj2qkkbil4n0wgn5cy0npz))/default.aspx',
        ('thor_gcadmin'):'https://gameadmin.12vin.com/',
        ('sta1_admin'):'https://admin.vina368.net/(S(eqjgza4zuwutd5y53y50m2w1))/default.aspx',
        ('sta1_agent'):'https://cmdbetagent.368aa.net/(S(s0oh1niwwjpjtpdfnu0l0r5b))/default.aspx',
        ('sta1_max222agent'):'https://max222agent.vina368.net/(S(ltsyfwcwxej34le5rrkoh2vf))/default.aspx',
        ('sta1_gcadmin'):'http://gcadmin.cmdbetsta.com/',
        ('sta2_admin'):'https://admin.cmmd368.com/(S(szdtumkbydibusat1o2xqqd1))/default.aspx',
        ('sta2_agent'):'https://cmdbetagent.cmmd368.com/(S(2ml21gvm0f0nwr1dkdqewr1j))/default.aspx',
        ('sta2_max222agent'):'https://max222agent.cmmd368.com/(S(s5qklauyhfsbuhe5qxlxcd02))/default.aspx',
        ('sta2_gcadmin'):'https://gcadmin.cmmd368.com/',
        ('prod_admin'):'https://admin.cmdbet.biz/(S(psil03jsmbweb24syq3pbbxi))/default.aspx',
        ('prod_agent'):'https://agent.cmdbet.com/(S(xk3u4trawghnzzifbcyd3pib))/default.aspx',
        ('prod_max222agent'):'https://agent.max222.com/(S(xdocse1gos51nz2rue5ao1ih))/default.aspx',
        ('prod_gcadmin'):'https://gcadmin.cmdbet.biz/'
    }
    
    url_key = f"{environment}_{website}"
    url = url_mapping.get(url_key)
    
    handlers = {
        "admin": {
            "thor": lambda: handle_thor(account, pswd, url),
            "sta1": lambda: handle_sta1(account, pswd, url),
            "sta2": lambda: handle_sta2(account, pswd, url),
            "prod": lambda: handle_prod(account, pswd, url),
        },
        "agent": {
            "thor": lambda: handle_thor(account, pswd, url),
            "sta1": lambda: handle_sta1(account, pswd, url),
            "sta2": lambda: handle_sta2(account, pswd, url),
            "prod": lambda: handle_prod(account, pswd, url),
        },
        "max222agent": {
            "thor": lambda: handle_thor(account, pswd, url),
            "sta1": lambda: handle_sta1(account, pswd, url),
            "sta2": lambda: handle_sta2(account, pswd, url),
            "prod": lambda: handle_prod(account, pswd, url),
        },
        "gcadmin": {
            "thor": lambda: handle_thor(account, pswd, url),
            "sta1": lambda: handle_sta1(account, pswd, url),
            "sta2": lambda: handle_sta2(account, pswd, url),
            "prod": lambda: handle_prod(account, pswd, url),
        }
    }
    
    handlers.get(website, {}).get(environment, default_handler)()


def default_handler():
    print('選擇的網站錯誤')   

def handle_thor(account, pswd, url):
    open_url(url)
    login(account, pswd)

def handle_sta1(account, pswd, url_mapping):
    print('agent')

def handle_sta2(account, pswd, url_mapping):
    print('max222agent')
    
def handle_prod(account, pswd, url_mapping):
    print('gcadmin')

def open_url(url):
    global chromeWeb
    chromeWeb = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install())
    )

    chromeWeb.maximize_window()
    chromeWeb.get(url)

def switch_to_frame(frame_name):
    """切換到指定的框架"""
    chromeWeb.switch_to.default_content()
    chromeWeb.switch_to.frame(frame_name)

def click_element_xpath(xpath):
    """等待並點擊指定的元素"""
    wait = WebDriverWait(chromeWeb, 10)  # 創建 WebDriverWait 對象
    result_click_element_xpath = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
    result_click_element_xpath.click()
    return 

def click_element_id(id):
    wait = WebDriverWait(chromeWeb, 10)  # 創建 WebDriverWait 對象
    result_click_element_id = wait.until(EC.element_to_be_clickable((By.ID, id)))
    result_click_element_id.click()
    return 

def report_game_name(xpath):
    select_outstanding_game_game = wait_chromeweb_xpath(xpath)
    time.sleep(1)
    select = Select(select_outstanding_game_game)
    options = select.options
    outstanding_admin_options_list = []
    for option in options:
        outstanding_admin_options_list.append(option.text)
    return outstanding_admin_options_list

def wait_chromeweb_id(id):
    waitChromewebId = WebDriverWait(chromeWeb, 10).until(
            EC.presence_of_element_located((By.ID, id))
    )
    return waitChromewebId
    
def wait_chromeweb_xpath(xpath):
    waitChromewebXpath = WebDriverWait(chromeWeb, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
    )
    return waitChromewebXpath

def login(account, pswd):
    # 取得驗證碼位置
    time.sleep(1)
    chromeWeb.save_screenshot("test.png")
    element = chromeWeb.find_element(By.ID, "verifyimg")
    verification_code = img(element)
    time.sleep(1)
    if verification_code != "111":
        while len(verification_code) != 4:
            chromeWeb.refresh()
            time.sleep(1)
            chromeWeb.get_screenshot_as_file("test.png")
            element = chromeWeb.find_element(By.ID, "verifyimg")
            verification_code = img(element)
    else:
        chromeWeb.quit()
        print("驗證碼解析失敗，請重新執行")
    # 登入開始
    chromeWeb.find_element(By.NAME, "UserName").send_keys(account)
    time.sleep(1)
    chromeWeb.find_element(By.NAME, "Password").send_keys(pswd)
    time.sleep(1)
    chromeWeb.find_element(By.XPATH, '//*[@id="txtInvalidation"]').send_keys(
        verification_code
    )
    print('輸入驗證碼完成')
    time.sleep(1)
    chromeWeb.find_element(By.NAME, "Submit").click() #登入按鈕


#驗證碼解析用
def img(element):
    try:
        element.location  # 取得圖片位置
        element.size  # 取得高度寬度
        left = element.location["x"]  # 取得左邊
        right = element.location["x"]+element.size["width"]  # 取得右邊
        top = element.location["y"]  # 取得上面
        bottom = element.location["y"]+element.size["height"]  # 取得下邊

        # 切割出驗證碼
        img_code = Image.open(".\\test.png")
        img_code.load()  # 確保圖片完全載入
        img_code = img_code.crop(
            (left, top, right, bottom))  # 順序一定要是 x,y,x+w,y+h
        img_code.save("verify.png", "png")

        # 對驗證碼進行處理
        img_code = img_code.convert("L")
        pix = img_code.load()
        w, h = img_code.size
        threshold = 205  # 畫素閾值

        # 遍歷所有畫素，大於閾值的為黑色
        for y in range(h):
            for x in range(w):
                if pix[x, y] < threshold:
                    pix[x, y] = 0
                else:
                    pix[x, y] = 255

        # 根據畫素二值結果重新生成圖片
        data = img_code.getdata()
        w, h = img_code.size
        black_point = 0
        for x in range(1, w - 1):
            for y in range(1, h - 1):
                mid_pixel = data[w * y + x]
                if mid_pixel < 50:
                    top_pixel = data[w * (y - 1) + x]
                    left_pixel = data[w * y + (x - 1)]
                    down_pixel = data[w * (y + 1) + x]
                    right_pixel = data[w * y + (x + 1)]
                    if top_pixel < 10:
                        black_point += 1
                    if left_pixel < 10:
                        black_point += 1
                    if down_pixel < 10:
                        black_point += 1
                    if right_pixel < 10:
                        black_point += 1
                    if black_point < 1:
                        img_code.putpixel((x, y), 255)
                    black_point = 0

        result = pytesseract.image_to_string(img_code)
        # 可能存在異常符號，用正則提取其中的數字
        regex = '\d+'
        result = ''.join(re.findall(regex, result))

        os.remove("verify.png")
        os.remove("test.png")
    except Exception as e:
        print(e)
        return '111'

    return result


def admin_function(request):
    
    #接收點選功能名稱
    action = request.POST.get('action')
    print(f'執行功能：{action}')
    #獲取當前url
    nowUrl = chromeWeb.current_url
    switch_to_frame("leftFrame")
    if action == 'AC Win Lose':
        acWinLoseReportresult = acWinLoseReport(nowUrl)
        print(acWinLoseReportresult)
        return HttpResponse(acWinLoseReportresult)
    elif action == 'Outstanding':
        outstandingReportresult = outstandingReport(nowUrl)
        print(outstandingReportresult)
        return HttpResponse(outstandingReportresult)
    elif action == 'Game Jackpot':
        gamejackreportresult = GameJackpotReport(nowUrl)
        print(gamejackreportresult)
        return HttpResponse(gamejackreportresult)
    elif action == 'Game Transaction':
        gametransactionresult = GameTransactionReport(nowUrl)
        print(gametransactionresult)
        return HttpResponse(gametransactionresult)
    elif action == 'Betlimit':
        betlimittresult = Betlimit(nowUrl)
        print(betlimittresult)
        return HttpResponse(betlimittresult)
    else:
        return action


def acWinLoseReport(nowUrl):
    if any(substring in nowUrl for substring in ['12vin', 'vina368', 'cmmd368']):
        admin_report_xpath = '//*[@id="div_leftLink"]/div[5]'
        #thor、sta1、sta2
        switch_to_frame("leftFrame")
        #等待Report可點擊
        click_element_xpath(admin_report_xpath)
        
        #等待AC WIN LOSE 可點擊
        try:
            click_element_xpath('//*[@id="div_leftLink"]/ul[5]/li[2]/a')
        except:
            #重新點report
            click_element_xpath(admin_report_xpath)
            #再點ac win lose
            click_element_xpath('//*[@id="div_leftLink"]/ul[5]/li[2]/a')
            
        #跳出leftframe   ，並進入mainFrame
        switch_to_frame("mainFrame")
        
        try:
            for i in range(1, 6):
                #從account win lose的SS層開始往下點
                click_element_xpath('//*[@id="tableGridView"]/tbody/tr[2]/td[1]/a')
                time.sleep(1)
        except:
            print("account win lose 目前無SS層資料")
        
        #casino
        click_element_xpath('//*[@id="form1"]/div[3]/div[1]/a[2]')

        time.sleep(1)
        #provider
        wait_chromeweb_id("slt_provider")
        provider =  chromeWeb.find_element(By.ID, "slt_provider")
        if game == "pp":
            Select(provider).select_by_value("22")  # 下拉選單取值 PP
        else:
            Select(provider).select_by_value("6")  # 下拉選單取值 SG
 
        #定位Game Name下拉選單定獲取下拉選單資料
        ac_win_lose_admin_list = report_game_name('//*[@id="slt_game"]')
        ac_num = 0
        gameComparisonResults = []
        for uploaded_file_name in gamename:
            ac_num += 1
            if ac_win_lose_admin_list.count(uploaded_file_name) == 1:  # x 在 admin_list出現次數是否為1
                gameComparisonResults.append(f'{str(ac_num)}. {uploaded_file_name}  --  PASS')
                print(f'{str(ac_num)}. {uploaded_file_name}  --  PASS')
                
            else:
                gameComparisonResults.append(f'{str(ac_num)}. {uploaded_file_name}  -- FAIL')
                print(f'{str(ac_num)}. {uploaded_file_name}  -- FAIL')
        print('AC WIN LOSE END')
        #跳出mainFrame，並進入leftFrame
        switch_to_frame("leftFrame")
        return gameComparisonResults
    
    else:
        pass
    
def outstandingReport(nowUrl):
    if any(substring in nowUrl for substring in ['12vin', 'vina368', 'cmmd368']):
        admin_report_xpath = '//*[@id="div_leftLink"]/div[5]'
        switch_to_frame('leftFrame')
        #點選Report
        time.sleep(1)
        try:
            click_element_xpath('//*[@id="div_leftLink"]/ul[5]/li[12]/a')
        except:
            click_element_xpath(admin_report_xpath)
            #點選outstanding
            click_element_xpath('//*[@id="div_leftLink"]/ul[5]/li[12]/a')
        switch_to_frame('mainFrame')
        try:
            for i in range(1, 6):
                #依序點到最後一層
                click_element_xpath('//*[@id="tableGridView"]/tbody/tr[2]/td[1]/a')
                time.sleep(1)
        except:
            print("outstanding 目前無SS層資料")
        #casino
        wait_chromeweb_id("slt_provider")
        provider = chromeWeb.find_element(By.ID, "slt_provider")
        if game == "pp":
            Select(provider).select_by_value("22")  # 下拉選單取值 PP
        else:
            Select(provider).select_by_value("6")  # 下拉選單取值 SG
            
        #定位Game Name下拉選單定獲取下拉選單資料
        outstanding_admin_options_list = report_game_name('//*[@id="slt_game"]')         
        out_num = 0
        for x in gamename:
            out_num += 1
            # x 在 admin_list出現次數是否為1
            if outstanding_admin_options_list.count(x) == 1:
                print(f'{str(out_num)}. {x}  -- PASS')
            else:
                print(f'{str(out_num)}. {x}  -- FAIL')  
        print('Outstanding END')
    else:
        pass
    
def GameJackpotReport(nowUrl):
    if any(substring in nowUrl for substring in ['12vin', 'vina368', 'cmmd368']):
        admin_report_xpath = '//*[@id="div_leftLink"]/div[5]'
        switch_to_frame('leftFrame')
        time.sleep(1)
        #點選Game Jack
        try:
            click_element_xpath('//*[@id="div_leftLink"]/ul[5]/li[17]/a')
        except:
            click_element_xpath(admin_report_xpath)
            time.sleep(1)
            click_element_xpath('//*[@id="div_leftLink"]/ul[5]/li[17]/a')
        switch_to_frame('mainFrame')
        
        wait_chromeweb_id('slt_provider')
        provider = chromeWeb.find_element(By.ID, 'slt_provider')
        if game == "PP":
            Select(provider).select_by_value("22")  # 下拉選單取值 PP
        else:
            Select(provider).select_by_value("6")  # 下拉選單取值 SG
        time.sleep(1)
        
        #定位Game Name下拉選單定獲取下拉選單資料
        Game_Jackpot_admin_options_list = report_game_name('//*[@id="slt_game"]')
        
        gj_num = 0
        for x in gamename:
            gj_num += 1
            # x 在 admin_list出現次數是否為1
            if Game_Jackpot_admin_options_list.count(x) == 1:
                print(f'{str(gj_num)}. {x}  -- PASS')
            else:
                print(f'{str(gj_num)}. {x}  -- FAIL')
        print('Game Jackpot END')
    else:
        pass
    
def GameTransactionReport(nowUrl):
    #thor、sta1、sta2
    if any(substring in nowUrl for substring in ['12vin', 'vina368', 'cmmd368']):
        switch_to_frame('leftFrame')
        #點選到Game Transaction
        time.sleep(1)
        try:
            click_element_xpath('//*[@id="div_leftLink"]/ul[13]/li[7]/a')
        except:
            click_element_xpath('//*[@id="div_leftLink"]/div[13]')
            time.sleep(0.5)
            click_element_xpath('//*[@id="div_leftLink"]/ul[13]/li[7]/a')
    #prod
    elif ('cmdbet' in nowUrl):
        switch_to_frame('leftFrame')
        #點選到Game Transaction
        try:
            time.sleep(1)
            click_element_xpath('//*[@id="div_leftLink"]/ul[12]/li[6]/a')
        except:
            time.sleep(1)
            click_element_xpath('//*[@id="div_leftLink"]/div[12]')
            click_element_xpath('//*[@id="div_leftLink"]/ul[12]/li[6]/a')
    else:
        return('url錯誤')
    
    switch_to_frame('mainFrame')
    wait_chromeweb_id("slt_provider")
    provider = chromeWeb.find_element(By.ID, "slt_provider")
    if game == "PP":
        Select(provider).select_by_value("22")  # 下拉選單取值 PP
    else:
        Select(provider).select_by_value("6")  # 下拉選單取值 SG
        
    #定位Game Name下拉選單定獲取下拉選單資料
    Game_Trancsaction_admin_options_list = report_game_name('//*[@id="slt_game"]')
    gt_num = 0
    for x in gamename:
        gt_num += 1
        # x 在 admin_list出現次數是否為1
        if Game_Trancsaction_admin_options_list.count(x) == 1:                                      
            print(f'{str(gt_num)}. {x}  -- PASS')
        else:
            print(f'{str(gt_num)}. {x}  -- FAIL')
    print('Game Transaction END')

def Betlimit(nowUrl):
    cur = {
            "AUD": "f1",
            "CNY": "f2",
            "EUR": "f4",
            "GBP": "f5",
            "HKD": "f6",
            "IDR": "f7",
            "JPY": "fy",
            "KRW": "fw",
            "MYR": "f3",
            "SGD": "fx",
            "USD": "fu",
            "VD": "fv",
            "INR": "fi",
            "BDT": "fo",
        }
    if environment == 'thor':
        cur['THB'] = 'fz'
    elif environment == 'sta1' or environment == 'sta2':
        cur['THB'] = 'ft'
    else:
        cur = {"USD": "f1"}
    if any(substring in nowUrl for substring in ['12vin', 'vina368', 'cmmd368']):
        try:
            click_element_xpath('//*[@id="div_leftLink"]/div[12]')
            time.sleep(0.5)
            click_element_xpath('//*[@id="div_leftLink"]/ul[12]/li[3]/a')
        except:
            switch_to_frame('leftFrame')
            click_element_xpath('//*[@id="div_leftLink"]/div[12]')
            time.sleep(0.5)
            click_element_xpath('//*[@id="div_leftLink"]/ul[12]/li[3]/a')
    else:
        try:
            #agent account list
            click_element_xpath('//*[@id="divLeftBox"]/div[6]')
            time.sleep(0.5)
            click_element_xpath('//*[@id="divLeftBox"]/ul[6]/li[2]/a')
        except:
            switch_to_frame('leftFrame')
            click_element_xpath('//*[@id="divLeftBox"]/div[6]')
            time.sleep(0.5)
            click_element_xpath('//*[@id="divLeftBox"]/ul[6]/li[2]/a')
    switch_to_frame('mainFrame')
    for key, value in cur.items():
        if environment != "PROD":
            try:
                time.sleep(3)
                select_cur = chromeWeb.find_element(By.ID, "slt_Currency")  # 定位幣別下拉選單
                Select(select_cur).select_by_value(key)  # 幣別下拉選單取值
            except:
                switch_to_frame("mainFrame")
                time.sleep(3)
                select_cur = chromeWeb.find_element(By.ID, "slt_Currency")  # 定位幣別下拉選單
                Select(select_cur).select_by_value(key)  # 幣別下拉選單取值
            time.sleep(0.5)
            chromeWeb.find_element(By.ID, "txt_UserName").clear()
            time.sleep(0.5)
            chromeWeb.find_element(By.ID, "txt_UserName").send_keys(value)
            time.sleep(0.5)
            # 定位submit並點選
            click_element_xpath('//*[@id="btn_submit"]')
            time.sleep(3)
            # 定位PP並點選
            click_element_xpath("//*[text()='[PP]']")
            time.sleep(2)
            chromeWeb.switch_to.window(chromeWeb.window_handles[1])  # 切換視窗
            time.sleep(0.5)
            try:
                # 定位Video Slots的Setting並點選
                click_element_xpath('//*[@id="tab_bettype"]/tbody/tr[5]/td/a') 
            except:
                chromeWeb.close()
                chromeWeb.switch_to.window(chromeWeb.window_handles[0])
                time.sleep(0.5)
                # 定位PP並點選
                click_element_xpath("//*[text()='[PP]']")
                chromeWeb.switch_to.window(chromeWeb.window_handles[1])  # 切換視窗
                # 定位Video Slots的Setting並點選
                click_element_xpath('//*[@id="tab_bettype"]/tbody/tr[5]/td/a')
            chromeWeb.maximize_window()  # 放大螢幕
            adminGameDict = {}  # admin字典
            # PP Video Stots頁數
            for x in range(0, 11):
                if x == 0:
                    adminGame = adminPage()
                    adminGameDict = dict(adminGameDict, **adminGame)  # 合併2個Dict
                else:
                    try:
                        ppVideoStotsPagexpath = f"/html/body/div[1]/div[3]/a[{x}]"
                        click_element_xpath(ppVideoStotsPagexpath)
                        time.sleep(1)
                        adminGame = adminPage()
                        adminGameDict = dict(adminGameDict, **adminGame)  # 合併2個Dict
                    except:
                        break

            chromeWeb.close()

            cur_betlimit_dict = {}
            pp_num = 0
            minRow = 0
            maxRow = 0
            for x in uploaded_file:
                file_data = x.read()
                excel = load_workbook(filename=BytesIO(file_data))  # 讀取excel
                for excelSheet in excel.sheetnames:
                    pp_num += 1
                    sheet = excel[excelSheet]  # 對EXCEL切換sheet
                    excelGame = sheet[1]
                    pp = excelGame[0].value
                    for x in range(1, 4):
                        # 判斷min和max
                        minMaxcolumn = sheet[2]
                        minMaxcolumnValue = minMaxcolumn[x].value
                        if minMaxcolumnValue[0:3] == "Min":
                            minRow = int(x)
                        elif minMaxcolumnValue[0:3] == "Max":
                            maxRow = int(x)
                        else:
                            pass
                    for column in range(3, 30):  # 對EXCEL的1~13行遍歷
                        cur_range = sheet[column]  # EXCEL該sheet的行數 column=行號
                        excke_cur = cur_range[0].value
                        if excke_cur == None:
                            break
                        else:
                            if key == "IDR" and excke_cur == "IDR2":
                                try:
                                    # excel文檔最小值 取小數後2位，沒有補0
                                    min = "%.2f" % cur_range[minRow].value
                                except:
                                    min = excelMinBetLimit(sheet, cur_range, minRow)
                                try:
                                    # excel文檔最大值 取小數後2位，沒有補0
                                    max = "%.2f" % cur_range[maxRow].value
                                except:
                                    max = excelMaxBetLimit(sheet, cur_range, maxRow)

                                betlimit = f'{min} ~ {max}'
                                # 加入cur_betlimit_dict字典
                                cur_betlimit_dict[pp] = betlimit
                                break
                            elif key == "VD" and excke_cur == "VND2":
                                try:
                                    min = (
                                        "%.2f" % cur_range[minRow].value
                                    )  # 抓取excel B欄位
                                except:
                                    min = excelMinBetLimit(sheet, cur_range, minRow)
                                try:
                                    max = (
                                        "%.2f" % cur_range[maxRow].value
                                    )  # 抓取excel C欄位
                                except:
                                    max = excelMaxBetLimit(sheet, cur_range, maxRow)
                                betlimit = f'{min} ~ {max}'
                                # 加入cur_betlimit_dict字典
                                cur_betlimit_dict[pp] = betlimit
                                break
                            elif excke_cur == key:
                                try:
                                    min = (
                                        "%.2f" % cur_range[minRow].value
                                    )  # 抓取excel B欄位
                                except:
                                    min = excelMinBetLimit(sheet, cur_range, minRow)
                                try:
                                    max = (
                                        "%.2f" % cur_range[maxRow].value
                                    )  # 抓取excel C欄位
                                except:
                                    max = excelMaxBetLimit(sheet, cur_range, maxRow)
                                betlimit = f'{min} ~ {max}'
                                # 加入cur_betlimit_dict字典
                                cur_betlimit_dict[pp] = betlimit
                                break
                    try:
                        # 比對excel的字典和admin字典的pp質是否相同
                        if cur_betlimit_dict[pp] == adminGameDict[pp] and pp != None:
                            print(f"{pp_num} :{pp} : {cur_betlimit_dict[pp]}(excel) / {adminGameDict[pp]}(admin) -- Pass ")
                        elif pp == None:
                            print(f"{pp_num}內容是空白的")
                        else:
                            print(f"{pp_num} :{pp} : {cur_betlimit_dict[pp]}(excel) / {adminGameDict[pp]}(admin) -- Fail")
                    except:
                        print(f"{pp_num} :{pp}無{key}資料")
    
def excelMinBetLimit(sheet, cur_range, minRow):
    excelbetLimit = cur_range[minRow].value
    excelbetLimit = excelbetLimit.replace("=", "")
    try:
        minValue = sheet[excelbetLimit].value
    except:
        excelbetLimitList = excelbetLimit.split("*")
        if isinstance(excelbetLimitList[0], int):
            minValue = sheet[excelbetLimitList[0]].value * int(excelbetLimitList[1])
        else:
            try:
                thisMin = sheet[excelbetLimitList[0]].value
            except:
                return 0
            try:
                thisMin = thisMin.replace("=", "")
                minValue = sheet[thisMin].value * int(excelbetLimitList[1])
            except:
                minValue = thisMin * int(excelbetLimitList[1])
    min = "%.2f" % minValue
    return min


def excelMaxBetLimit(sheet, cur_range, maxRow):
    excelbetLimit = cur_range[maxRow].value
    excelbetLimit = excelbetLimit.replace("=", "")
    try:
        maxValue = sheet[excelbetLimit].value
    except:
        excelbetLimitList = excelbetLimit.split("*")
        if isinstance(excelbetLimitList[0], int):
            maxValue = sheet[excelbetLimitList[0]].value * int(excelbetLimitList[1])
        else:
            try:
                thisMax = sheet[excelbetLimitList[0]].value
            except:
                return 0
            try:
                thisMax = thisMax.replace("=", "")
                maxValue = sheet[thisMax].value * int(excelbetLimitList[1])
            except:
                maxValue = thisMax * int(excelbetLimitList[1])
    max = "%.2f" % maxValue
    return max
 
def adminPage():
    cur_dict = {}
    cur_dict.clear()
    for num in range(1, 101):  # 遍歷第一頁1~100行，取Game Name 和 Bet Limit
        try:
            gameName1 = chromeWeb.find_element(
                By.XPATH, f'//*[@id="tablelist"]/tbody/tr[{num}]/td[2]'
            ).text
            betLimit1 = chromeWeb.find_element(
                By.XPATH, f'//*[@id="tablelist"]/tbody/tr[{num}]/td[3]'
            ).text
        except:
            break
        cur_dict[gameName1] = betLimit1  # 取出職加入admin字典

    return cur_dict

def oddsConversion(requset):
    return render(requset, "oddsConversion.html")