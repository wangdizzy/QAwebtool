from django.shortcuts import render
from django.shortcuts import render
from datetime import datetime
from home.models import Post
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from django.http import JsonResponse
from oauth2client.service_account import ServiceAccountCredentials
from home.views import chromeWeb
from home import views as home_views
from urllib.parse import urlparse
from io import BytesIO
import time
import gspread
import pandas as pd


def google_sheet(worksheet_name="gameData(PP)"):

    sheet_json = {
        "type": "service_account",
        "project_id": "inlaid-sentinel-407402",
        "private_key_id": "e8289caa403c29ec8edeb62490cdb696984dea98",
        "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQC0FOWiThPW8PVC\nHxJT6zFuXVE5ekvsIDiDzznJRorYkZFRxV6n6H2c6rBUZQf2dhgYR6HorFcMz+oy\n9ihlO2Fv7fXD6bfn7xje56JPyVu3RXsmAB1K7eZYFjWYeULYe3Xl5eLZX4QYFSvI\nBK/CX+pboojywnPcJlSBxqW1q9cz+crvpfTXq5mPEdp1sdotOxmX6pTi8SvB4eSd\n3A31oCtpsEvXFwK1ZUwEVlw7cLdgGqhKrEgn8wHVrYiVYSA7pM19opy2gJJyqZVX\n6wzeGA1uBGTXDAEMBLJ8AStuFAjoG9/3OMay0Ln3ba9lc7vOWsdBFz/TqdSaRGre\nUJuDuFvJAgMBAAECggEAEfQOY8TIdhKeRPoPi/XD/xaHGUWfyZn5wGxZvL8PayFy\nHTah5fgIA+ui6jsLVO83nj4P/oAmCpU06mE/rD4EDBJrgN1tdA5SirCJrk4rGmWv\nLh3vTa/Tmd8W901JlIcUfTfSyqya4QLFU2LiOezxkrKs2BT6U42vuaN6FFdeNGRt\nPLPxSnNHwpv59XrCfFoOp78oU3LrE+J7DQrjITBSy3iVZQypW+x5JapJ9R5Ix5Ox\njHehNaqTJ4K5y31ZBEur2ZQTNEd95i+gBIHAJbb3L49C2zt0TwFbEVkdOeG5zZ3l\nJr9oPRYQN5QwlgZgIW8T5SKlRA8rsv7wAwm9qV/XXQKBgQDoVrQiTafWuLD1VHxJ\n3uMsI+dzMFRBKGmd8tbFD5V1atFrnXSTVBtg2RdsAM21WIG8D0EhLD5Kc7XlpIPv\n9sMWEaKKbGLk8a4eKXAEetsJjnnCdCLb+GqeO2Qp0JDZ6cKwZFKNopcDEEiEX39i\nZtHL82Eq7DbnkLJ0G8dwqZJB7QKBgQDGa80FkScVnyw5G0PEqtQ1wAfytJyGB3V6\nJDNdaC8XXm2vzUHWZM/Tyrw+qeHZRuLDu2gb2p221zE7CyYkxgj66VWBNRPnpEiY\nc3VV7bXBfVEAiftVRWAHgrHtGCRACKKt4OgfrJrzRAcekye4Lm4er4QDEZI4uBZq\nYm8fqVG1zQKBgQDegDoeJ9Q2M8V0DKbCb6uK2A+NJplplPQgiTDYo2X0folzz+SW\nOxPFGeHuUo6tvsbvfIRY6m/1CP8HnxejNOP7PIQ2oDnNGw4uYGygPa+KZWGBsYq4\nshwY0LPJv60Yo18JYeoVLcIE8xEfg/0QFXuRH9DMNE8YUGA2BWxoHlysuQKBgQCl\nRMcpGuTeGo1gJ3iDx/InrwIvwwYYkP/ls19hLtUCdvGPm7x50dBVTSkMXL20F1nr\nxB4MDUSONaFY14l22chDDbTdgRNKPskEyi5yWyOnvTSJ6WQBe15oAxEmNZSEDW1K\nvOk68K7DbucrLVDJFUs9nd2sHKeZPKPXCpQaYBKiBQKBgEdAQVOL6p5sm/acvKsF\nSzzieDOzCeeiGYmeqguiEg7sLJ5UcSlCdKDGKFtgc1vi9j1cVcAah2PBgigyuLhx\nrkHfPNKePKZ/Do3WTuG2HOmlRXK/FJXsE+9m6joD9ZVipE2mTowWNq675EkUkfrd\n/Ef2Q0gOIPY2sY5/aKQKR18P\n-----END PRIVATE KEY-----\n",
        "client_email": "qa-451@inlaid-sentinel-407402.iam.gserviceaccount.com",
        "client_id": "115465534078375635707",
        "auth_uri": "https://accounts.google.com/o/oauth2/auth",
        "token_uri": "https://oauth2.googleapis.com/token",
        "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
        "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/qa-451%40inlaid-sentinel-407402.iam.gserviceaccount.com",
        "universe_domain": "googleapis.com",
    }

    # 定義存取的範圍 feeds = google sheet
    scopes = ["https://spreadsheets.google.com/feeds"]

    # 將金鑰放在dict使用
    credentials = ServiceAccountCredentials.from_json_keyfile_dict(sheet_json, scopes)

    # 傳入gspread模組
    client = gspread.authorize(credentials)

    # 使用open_by_key方式傳入google sheet 金鑰
    spreadsheet = client.open_by_key("1qWdc0QTGY13LEsr_N_5SA4cfyWTHaGXk8GhJFosoNzc")

    # 指定執行頁面
    gameData = spreadsheet.worksheet(worksheet_name)

    return gameData

def google_betlimit_data(currency):
    """
    解析google sheet內currency的betlimit資料，並回傳一個字典，格式為{game: betlimit}
    """
    gamename = google_sheet()
    
    #取得所有資料
    all_data = gamename.get_all_values()
    
    #轉換成 DataFrame並跳過第一行標題
    df = pd.DataFrame(all_data[1:], columns=all_data[0])
    
    # 只保留需要的欄位
    df = df[['Game', 'Currency', 'MinMax']]
    
    # 過濾出指定貨幣的資料
    dfGameMinMax = df[df['Currency'] == currency].copy()
    
    #儲存資料字典
    gameMinMaxDict = dfGameMinMax.set_index('Game')['MinMax'].to_dict()

    return gameMinMaxDict
    
'''
def excel_file_name(file):
    gamename = []
    for uploaded_file_name in file:
            excel_game_name = uploaded_file_name.name
            excel_game_name = excel_game_name.replace('.xlsx','')
            gamename.append(excel_game_name)
    return gamename
'''

def ppsgFunctionSelection(request):
    posts = Post.objects.all()
    now = datetime.now()
    return render(request, "ppsgFunctionSelection.html", locals())

'''
def excel_list(file_path):
    gameName_list = []
    # excel = load_workbook(file_path)  # 讀取excel檔
    for x in file_path:
        excel = load_workbook(x)
        for sheetName in excel.sheetnames:  # 把excel 的sheet全部取出遍歷
            try:
                sheet = excel[sheetName]  # 對EXCEL切換sheet
            except:
                break
            excelGame = sheet[1]
            excelGameName = excelGame[0].value
            if excelGameName is not None:
                gameNameRow = sheet[1]
                gameNameValue = gameNameRow[0].value
                gameName_list.append(gameNameValue)
    return gameName_list
'''
'''
def excel_gcadmin(file_path):
    gameTypeGameName = {}
    # excel = load_workbook(file_path)
    excel = load_workbook(file_path)
    sheet = excel["Sheet1"]
    for row in range(2, 100):
        gametype_row = sheet[row]
        gametype_value = gametype_row[0].value
        if gametype_value is not None:
            ProviderGameType = gametype_row[1].value
            gameNameEnglishValue = gametype_row[2].value
            gameTypeGameName[gametype_value] = (
                gameNameEnglishValue,
                ProviderGameType,
            )

    return gameTypeGameName
'''
def switch_to_frame(frame_name):
    """切換到指定的框架"""
    home_views.chromeWeb.switch_to.default_content()
    home_views.chromeWeb.switch_to.frame(frame_name)

def click_element_xpath(xpath):
    """等待並點擊指定的元素"""
    try:
        wait = WebDriverWait(home_views.chromeWeb, 10)  # 創建 WebDriverWait 對象
        result_click_element_xpath = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
        time.sleep(1)
        result_click_element_xpath.click()
        return True
    except:
        print(f"點擊[{xpath}]失敗")
        return False

def click_element_id(id):
    try:
        wait = WebDriverWait(home_views.chromeWeb, 10)  # 創建 WebDriverWait 對象
        result_click_element_id = wait.until(EC.element_to_be_clickable((By.ID, id)))
        time.sleep(1)
        result_click_element_id.click()
        return True
    except:
        print(f"點擊[{id}]失敗")
        return False

def report_game_name(xpath):
    time.sleep(1)
    select_outstanding_game_game = wait_chromeweb_xpath(xpath)
    time.sleep(1)
    select = Select(select_outstanding_game_game)
    time.sleep(1)
    options = select.options
    outstanding_admin_options_list = []
    for option in options:
        outstanding_admin_options_list.append(option.text)
    return outstanding_admin_options_list

def wait_chromeweb_id(id):
    waitChromewebId = WebDriverWait(home_views.chromeWeb, 10).until(
            EC.presence_of_element_located((By.ID, id))
    )
    return waitChromewebId
    
def wait_chromeweb_xpath(xpath):
    waitChromewebXpath = WebDriverWait(home_views.chromeWeb, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
    )
    return waitChromewebXpath

def report_function(request):
    global get_sheet_a_game_name, game, environment, get_sheet_a_game_name
    
    #從session獲取資料
    environment = request.session.get('environment')
    game = request.session.get('game')
    #接收點選功能名稱
    action = request.POST.get('action')
    print(f'執行功能：{action}')
    #使用 home.views 的全域變數，獲取當前url
    nowUrl = home_views.chromeWeb.current_url
    gamename = google_sheet()
    #取得工作表的第 1 欄（A 欄）的所有值，透過[1:]跳過第一行標題
    get_sheet_a_game_name = gamename.col_values(1)[1:]
    
    #自動去除重複的元素，並轉成list
    get_sheet_a_game_name = list(set(get_sheet_a_game_name))
    switch_to_frame("leftFrame")
    if action == 'AC Win Lose':
        acWinLoseReportresult = acWinLoseReport(nowUrl)
        print(acWinLoseReportresult)
        return JsonResponse({
            'success': True,
            'action': action,
            'result': acWinLoseReportresult
        })
    elif action == 'Outstanding':
        outstandingReportresult = outstandingReport(nowUrl)
        print(outstandingReportresult)
        return JsonResponse({
            'success': True,
            'action': action,
            'result': outstandingReportresult
        })
    elif action == 'Game Jackpot':
        gamejackreportresult = GameJackpotReport(nowUrl)
        print(gamejackreportresult)
        return JsonResponse({
            'success': True,
            'action': action,
            'result': gamejackreportresult
        })
    elif action == 'Game Transaction':
        gametransactionresult = GameTransactionReport(nowUrl)
        print(gametransactionresult)
        return JsonResponse({
            'success': True,
            'action': action,
            'result': gametransactionresult
        })
    elif action == 'Betlimit':
        betlimittresult = Betlimit(nowUrl)
        print(betlimittresult)
        return JsonResponse({
            'success': True,
            'action': action,
            'result': betlimittresult
        })
    elif action == 'Game List':
        betlimittresult = GameList(nowUrl)
        print(betlimittresult)
        return JsonResponse({
            'success': True,
            'action': action,
            'result': betlimittresult
        })
    else:
        return action


def acWinLoseReport(nowUrl):
    hostname = urlparse(nowUrl).hostname
    if any(substring in nowUrl for substring in ['12vin', 'vina368', 'cmmd368', '368aa']):
        '''
            thor admin  = 12vin / sta admin = vina368 / sta2 admin = cmmd368
            thor anget = 12vin / sta agent = 368aa、vina368 / sta2 agent = cmmd368
        '''
        admin_report_xpath = '//*[@id="div_leftLink"]/div[5]'
        admin_acwinlose_xpath = '//*[@id="div_leftLink"]/ul[5]/li[2]/a'
        anget_report_xpath = '//*[@id="divLeftBox"]/div[5]'
        anget_acwinlose_xpath = '//*[@id="divLeftBox"]/ul[5]/li[2]/a' 
        max222anget_report_xpath = '//*[@id="divLeftBox"]/div[5]'
        max222anget_acwinlose_xpath = '//*[@id="divLeftBox"]/ul[5]/li[3]/a' 
        
        switch_to_frame("leftFrame")
        
        #等待AC WIN LOSE 可點擊
        if hostname.startswith('max222agent.12vin'):
            print('進入max222agent')
            if not click_element_xpath(max222anget_acwinlose_xpath):
                time.sleep(1)
                #重新點report
                click_element_xpath(max222anget_report_xpath)
                #再點outstanding
                time.sleep(1)
                click_element_xpath(max222anget_acwinlose_xpath)  
        elif hostname.startswith('agent.12vin'):
            print('進入agent')
            if not click_element_xpath(anget_acwinlose_xpath):
                #重新點report
                click_element_xpath(anget_report_xpath)
                #再點outstanding
                time.sleep(3)
                click_element_xpath(anget_acwinlose_xpath)    
        else:
            print('進入admin')
            if not click_element_xpath(admin_acwinlose_xpath):
                #重新點report
                click_element_xpath(admin_report_xpath)
                #再點outstanding
                time.sleep(1)
                click_element_xpath(admin_acwinlose_xpath)
            
        #跳出leftframe   ，並進入mainFrame
        switch_to_frame("mainFrame")
        
        for i in range(1, 6):
            #從account win lose的SS層開始往下點
            try:
                time.sleep(1)
                click_element_xpath('//*[@id="tableGridView"]/tbody/tr[2]/td[1]/a')
                time.sleep(1)
            except:
                break
        #casino
        click_element_xpath('//*[@id="form1"]/div[3]/div[1]/a[2]')

        time.sleep(1)
        #provider
        wait_chromeweb_id("slt_provider")
        provider =  home_views.chromeWeb
        provider = provider.find_element(By.ID, "slt_provider")
        
        if game == "pp":
            Select(provider).select_by_value("22")  # 下拉選單取值 PP
        else:
            Select(provider).select_by_value("6")  # 下拉選單取值 SG
 
        #定位Game Name下拉選單定獲取下拉選單資料
        ac_win_lose_admin_list = report_game_name('//*[@id="slt_game"]')
        ac_num = 0
        gameComparisonResults = []
        for uploaded_file_name in get_sheet_a_game_name:
            ac_num += 1
            if ac_win_lose_admin_list.count(uploaded_file_name) == 1:  # x 在 admin_list出現次數是否為1
                gameComparisonResults.append(f'{str(ac_num)}. {uploaded_file_name}  --  PASS')
                
            else:
                gameComparisonResults.append(f'{str(ac_num)}. {uploaded_file_name}  -- FAIL')
                
        print('AC WIN LOSE END')
        #跳出mainFrame，並進入leftFrame
        switch_to_frame("leftFrame")
        return gameComparisonResults
    
    else:
        pass
    
def outstandingReport(nowUrl):
    hostname = urlparse(nowUrl).hostname
    if any(substring in nowUrl for substring in ['12vin', 'vina368', 'cmmd368']):
        '''
            thor admin  = 12vin / sta admin = vina368 / sta2 admin = cmmd368
            thor anget = 12vin / sta agent = 368aa、vina368 / sta2 agent = cmmd368
        ''' 
        admin_report_xpath = '//*[@id="div_leftLink"]/div[5]'
        admin_outstanding_xpath = '//*[@id="div_leftLink"]/ul[5]/li[12]/a'
        anget_report_xpath = '//*[@id="divLeftBox"]/div[5]'
        anget_outstanding_xpath = '//*[@id="divLeftBox"]/ul[5]/li[12]/a' 
        max222anget_report_xpath = '//*[@id="divLeftBox"]/div[5]/a'
        max222anget_outstanding_xpath = '//*[@id="divLeftBox"]/ul[5]/li[13]/a' 
        
        switch_to_frame('leftFrame')
        #點選Report
        time.sleep(1)
        if hostname.startswith('max222agent.12vin'):
            print('進入max222agent')
            if not click_element_xpath(max222anget_outstanding_xpath):
                time.sleep(1)
                #重新點report
                click_element_xpath(max222anget_report_xpath)
                #再點outstanding
                time.sleep(1)
                click_element_xpath(max222anget_outstanding_xpath)  
        elif hostname.startswith('agent.12vin'):
            print('進入agent')
            if not click_element_xpath(anget_outstanding_xpath):
                #重新點report
                click_element_xpath(anget_report_xpath)
                #再點outstanding
                time.sleep(3)
                click_element_xpath(anget_outstanding_xpath)    
        else:
            print('進入admin')
            if not click_element_xpath(admin_outstanding_xpath):
                #重新點report
                click_element_xpath(admin_report_xpath)
                #再點outstanding
                time.sleep(1)
                click_element_xpath(admin_outstanding_xpath)
            
        switch_to_frame('mainFrame')
        
        for i in range(1, 6):
            try:
                #依序點到最後一層
                time.sleep(1)
                click_element_xpath('//*[@id="tableGridView"]/tbody/tr[2]/td[1]/a')
                time.sleep(1)
            except:
                break

        #casino   
        click_element_xpath('//*[@id="form1"]/div[3]/div[1]/a[2]')
        
        #provider
        wait_chromeweb_id("slt_provider")
        provider = home_views.chromeWeb
        provider = provider.find_element(By.ID, "slt_provider")
        if game == "pp":
            Select(provider).select_by_value("22")  # 下拉選單取值 PP
        else:
            Select(provider).select_by_value("6")  # 下拉選單取值 SG
            
        #定位Game Name下拉選單定獲取下拉選單資料
        outstanding_admin_options_list = report_game_name('//*[@id="slt_game"]')         
        out_num = 0
        outstandingReportresult = []
        for x in get_sheet_a_game_name:
            out_num += 1
            # x 在 admin_list出現次數是否為1
            if outstanding_admin_options_list.count(x) == 1:
                outstandingReportresult.append(f'{str(out_num)}. {x}  --  PASS')
            else:
                outstandingReportresult.append(f'{str(out_num)}. {x}  --  FAIL')
        print('Outstanding END')
        switch_to_frame("leftFrame")
        return outstandingReportresult
    else:
        pass
    
def GameJackpotReport(nowUrl):
    if any(substring in nowUrl for substring in ['12vin', 'vina368', 'cmmd368']):
        switch_to_frame('leftFrame')
        time.sleep(1)
        #點選Game Jack
        if not click_element_xpath('//*[@id="div_leftLink"]/ul[5]/li[17]/a'):
            
            click_element_xpath('//*[@id="div_leftLink"]/div[5]')
            time.sleep(1)
            click_element_xpath('//*[@id="div_leftLink"]/ul[5]/li[17]/a')
        
        switch_to_frame('mainFrame')
        
        time.sleep(1)
        wait_chromeweb_id('slt_provider')
        provider = home_views.chromeWeb
        provider = provider.find_element(By.ID, "slt_provider")
        if game == "pp":
            Select(provider).select_by_value("22")  # 下拉選單取值 PP
        else:
            Select(provider).select_by_value("6")  # 下拉選單取值 SG
        time.sleep(1)
        
        #定位Game Name下拉選單定獲取下拉選單資料
        Game_Jackpot_admin_options_list = report_game_name('//*[@id="slt_game"]')
        
        gj_num = 0
        GameJackpotReportresult = []
        for x in get_sheet_a_game_name:
            gj_num += 1
            # x 在 admin_list出現次數是否為1
            if Game_Jackpot_admin_options_list.count(x) == 1:
                GameJackpotReportresult.append(f'{str(gj_num)}. {x}  --  PASS')
            else:
                GameJackpotReportresult.append(f'{str(gj_num)}. {x}  --  FAIL')
        print('Game Jackpot END')
        switch_to_frame("leftFrame")
        return GameJackpotReportresult
    else:
        pass
    
def GameTransactionReport(nowUrl):
    #thor、sta1、sta2
    if any(substring in nowUrl for substring in ['12vin', 'vina368', 'cmmd368']):
        switch_to_frame('leftFrame')
        #點選到Game Transaction
        time.sleep(1)
        if not click_element_xpath('//*[@id="div_leftLink"]/ul[13]/li[7]/a'):
            time.sleep(1)
            click_element_xpath('//*[@id="div_leftLink"]/div[13]')
            time.sleep(1)
            click_element_xpath('//*[@id="div_leftLink"]/ul[13]/li[7]/a')
    #prod
    elif 'cmdbet' in nowUrl:
        switch_to_frame('leftFrame')
        #點選到Game Transaction
        try:
            time.sleep(1)
            click_element_xpath('//*[@id="div_leftLink"]/ul[12]/li[6]/a')
        except:
            time.sleep(1)
            click_element_xpath('//*[@id="div_leftLink"]/div[12]')
            click_element_xpath('//*[@id="div_leftLink"]/ul[12]/li[6]/a')
    else:
        return('url錯誤')
    time.sleep(1)
    
    switch_to_frame('mainFrame')
    wait_chromeweb_id("slt_provider")
    provider = home_views.chromeWeb
    provider = provider.find_element(By.ID, "slt_provider")
    if game == "pp":
        Select(provider).select_by_value("22")  # 下拉選單取值 PP
    else:
        Select(provider).select_by_value("6")  # 下拉選單取值 SG
        
    #定位Game Name下拉選單定獲取下拉選單資料
    Game_Trancsaction_admin_options_list = report_game_name('//*[@id="slt_game"]')
    gt_num = 0
    GameTransactionReportresult = []
    for x in get_sheet_a_game_name:
        gt_num += 1
        # x 在 admin_list出現次數是否為1
        if Game_Trancsaction_admin_options_list.count(x) == 1:                                      
            GameTransactionReportresult.append(f'{str(gt_num)}. {x}  --  PASS')
        else:
            GameTransactionReportresult.append(f'{str(gt_num)}. {x}  --  FAIL')
    print('Game Transaction END')
    switch_to_frame("leftFrame")
    return GameTransactionReportresult

def Betlimit(nowUrl):
    cur = {
        "AUD": "f1",
        "CNY": "f2",
        "EUR": "f4",
        "GBP": "f5",
        "HKD": "f6",
        "IDR": "f7",
        "JPY": "fy",
        "KRW": "fw",
        "MYR": "f3",
        "SGD": "fx",
        "USD": "fu",
        "VD": "fv",
        "INR": "fi",
        "BDT": "fo",
    }
    
    if environment == 'thor':
        cur['THB'] = 'fz'
    elif environment == 'sta1' or environment == 'sta2':
        cur['THB'] = 'ft'
    else:
        cur = {"USD": "f1"}
    if any(substring in nowUrl for substring in ['12vin', 'vina368', 'cmmd368']):
        if click_element_xpath('//*[@id="div_leftLink"]/div[12]'):
            time.sleep(0.5)
            click_element_xpath('//*[@id="div_leftLink"]/ul[12]/li[3]/a')
        else:
            switch_to_frame('leftFrame')
            click_element_xpath('//*[@id="div_leftLink"]/div[12]')
            time.sleep(0.5)
            click_element_xpath('//*[@id="div_leftLink"]/ul[12]/li[3]/a')
    else:
        #agent account list
        if click_element_xpath('//*[@id="divLeftBox"]/div[6]'):
            time.sleep(1)
            click_element_xpath('//*[@id="divLeftBox"]/ul[6]/li[2]/a')
        else:
            switch_to_frame('leftFrame')
            click_element_xpath('//*[@id="divLeftBox"]/div[6]')
            time.sleep(0.5)
            click_element_xpath('//*[@id="divLeftBox"]/ul[6]/li[2]/a')
            
    switch_to_frame('mainFrame')
    n = 1
    betlimitResult = []
    for key, value in cur.items():
        n += 1
        if environment != "PROD":
            try:
                time.sleep(3)
                select_cur = home_views.chromeWeb.find_element(By.ID, "slt_Currency")  # 定位幣別下拉選單
                Select(select_cur).select_by_value(key)  # 幣別下拉選單取值
            except:
                switch_to_frame("mainFrame")
                time.sleep(3)
                select_cur = home_views.chromeWeb.find_element(By.ID, "slt_Currency")  # 定位幣別下拉選單
                Select(select_cur).select_by_value(key) # 幣別下拉選單取值
            time.sleep(1)
            home_views.chromeWeb.find_element(By.ID, "txt_UserName").clear()
            time.sleep(1)
            home_views.chromeWeb.find_element(By.ID, "txt_UserName").send_keys(value)
            print('Account List點輸入ss完成')
            time.sleep(1)
            
            # 定位submit並點選
            click_element_xpath('//*[@id="btn_submit"]')
            print('Account List點submit完成')
            time.sleep(3)
            
            # 定位PP並點選
            click_element_xpath("//*[text()='[PP]']")
            print('Account List點PP完成')
            time.sleep(2)
            home_views.chromeWeb.switch_to.window(home_views.chromeWeb.window_handles[1])  # 切換視窗
            time.sleep(0.5)
            try:
                # 定位Video Slots的Setting並點選
                click_element_xpath('//*[@id="tab_bettype"]/tbody/tr[5]/td/a') 
            except:
                home_views.chromeWeb.close()
                home_views.chromeWeb.switch_to.window(home_views.chromeWeb.window_handles[0])
                time.sleep(1)
                
                # 定位PP並點選
                click_element_xpath("//*[text()='[PP]']")
                home_views.chromeWeb.switch_to.window(home_views.chromeWeb.window_handles[1])  # 切換視窗
                # 定位Video Slots的Setting並點選
                click_element_xpath('//*[@id="tab_bettype"]/tbody/tr[5]/td/a')
            
            # 放大螢幕
            home_views.chromeWeb.maximize_window()  
            
            # admin字典
            adminGameDict = {}  
            
            # PP Video Stots頁數
            for x in range(0, 11):
                if x == 0:
                    adminGame = adminPage()
                    adminGameDict = dict(adminGameDict, **adminGame)  # 合併2個Dict
                else:
                    try:
                        ppVideoStotsPagexpath = f"/html/body/div[1]/div[3]/a[{x}]"
                        if (click_element_xpath(ppVideoStotsPagexpath)) is True :
                            time.sleep(1)
                            adminGame = adminPage()
                            adminGameDict = dict(adminGameDict, **adminGame)  # 合併2個Dict
                        else:
                            break
                    except:
                        print(f'沒有第{x}頁了')
                        break
            
            home_views.chromeWeb.close()
            home_views.chromeWeb.switch_to.window(home_views.chromeWeb.window_handles[0])
            time.sleep(1)

            cur_betlimit_dict = google_betlimit_data(key)
            pp_num = 0 
            print('開始比對Betlimit')
            print(f'cur_betlimit_dict: {cur_betlimit_dict}')
            # 比對excel的字典和admin字典的pp質是否相同
            try:
                for gameName in get_sheet_a_game_name: 
                    pp_num += 1
                    if cur_betlimit_dict[gameName] == adminGameDict[gameName] and gameName != None:
                        betlimitResult.append(f'{key}--{pp_num} :{gameName} : {cur_betlimit_dict[gameName]}(excel) / {adminGameDict[gameName]}(admin) -- Pass')
                    elif gameName == None:
                        betlimitResult.append(f"{key}--{gameName}內容是空白的")
                    else:
                        betlimitResult.append(f"{key}--{pp_num} :{gameName} : {cur_betlimit_dict[gameName]}(excel) / {adminGameDict[gameName]}(admin) -- Fail")
            except Exception as e:
                print(f'發生錯誤遊戲：{key}--{gameName}, 錯誤訊息：{e}')
            #for x in uploaded_file:
            #    file_data = x.read()
            #    excel = load_workbook(filename=BytesIO(file_data))  # 讀取excel
            #    for excelSheet in excel.sheetnames:
            #        pp_num += 1
            #        sheet = excel[excelSheet]  # 對EXCEL切換sheet
            #        excelGame = sheet[1]
            #        pp = excelGame[0].value
            #        for x in range(1, 4):
            #            # 判斷min和max
            #            minMaxcolumn = sheet[2]
            #            minMaxcolumnValue = minMaxcolumn[x].value
            #            if minMaxcolumnValue[0:3] == "Min":
            #                minRow = int(x)
            #            elif minMaxcolumnValue[0:3] == "Max":
            #                maxRow = int(x)
            #            else:
            #                pass
            #        for column in range(3, 30):  # 對EXCEL的1~13行遍歷
            #            cur_range = sheet[column]  # EXCEL該sheet的行數 column=行號
            #            excke_cur = cur_range[0].value
            #            if excke_cur == None:
            #                break
            #            else:
            #                if key == "IDR" and excke_cur == "IDR2":
            #                    try:
            #                        # excel文檔最小值 取小數後2位，沒有補0
            #                        min = "%.2f" % cur_range[minRow].value
            #                    except:
            #                        min = excelMinBetLimit(sheet, cur_range, minRow)
            #                    try:
            #                        # excel文檔最大值 取小數後2位，沒有補0
            #                        max = "%.2f" % cur_range[maxRow].value
            #                    except:
            #                        max = excelMaxBetLimit(sheet, cur_range, maxRow)
#
            #                    betlimit = f'{min} ~ {max}'
            #                    # 加入cur_betlimit_dict字典
            #                    cur_betlimit_dict[pp] = betlimit
            #                    break
            #                elif key == "VD" and excke_cur == "VND2":
            #                    try:
            #                        min = (
            #                            "%.2f" % cur_range[minRow].value
            #                        )  # 抓取excel B欄位
            #                    except:
            #                        min = excelMinBetLimit(sheet, cur_range, minRow)
            #                    try:
            #                        max = (
            #                            "%.2f" % cur_range[maxRow].value
            #                        )  # 抓取excel C欄位
            #                    except:
            #                        max = excelMaxBetLimit(sheet, cur_range, maxRow)
            #                    betlimit = f'{min} ~ {max}'
            #                    # 加入cur_betlimit_dict字典
            #                    cur_betlimit_dict[pp] = betlimit
            #                    break
            #                elif excke_cur == key:
            #                    try:
            #                        min = (
            #                            "%.2f" % cur_range[minRow].value
            #                        )  # 抓取excel B欄位
            #                    except:
            #                        min = excelMinBetLimit(sheet, cur_range, minRow)
            #                    try:
            #                        max = (
            #                            "%.2f" % cur_range[maxRow].value
            #                        )  # 抓取excel C欄位
            #                    except:
            #                        max = excelMaxBetLimit(sheet, cur_range, maxRow)
            #                    betlimit = f'{min} ~ {max}'
            #                    # 加入cur_betlimit_dict字典
            #                    cur_betlimit_dict[pp] = betlimit
            #                    break
            #        try:
            #            # 比對excel的字典和admin字典的pp質是否相同
            #            betlimitResult = []
            #            if cur_betlimit_dict[pp] == adminGameDict[pp] and pp != None:
            #                betlimitResult.append(f'{pp_num} :{pp} : {cur_betlimit_dict[pp]}(excel) / {adminGameDict[pp]}(admin) -- Pass')
            #            elif pp == None:
            #                betlimitResult.append(f"{pp_num}內容是空白的")
            #            else:
            #                betlimitResult.append(f"{pp_num} :{pp} : {cur_betlimit_dict[pp]}(excel) / {adminGameDict[pp]}(admin) -- Fail")
            #        except:
            #            betlimitResult.append(f"{pp_num} :{pp} 在admin沒有資料")
    return betlimitResult

def GameList(nowUrl):
    #建立空陣列
    valueList = []
    
    sheet = google_sheet('GameType')
    
    if any(substring in nowUrl for substring in ['12vin', 'vina368', 'cmmd368']):
        gcadmin_general_xpath = '/html/body/div/table/tbody/tr/td[1]/div/div[1]'
        game_list_xpath = '/html/body/div/table/tbody/tr/td[1]/div/ul[1]/li[4]/a'
        game_gameList_submit_xpath = '//*[@id="form1"]/input[2]'
        #thor、sta1、sta2
        switch_to_frame("leftFrame")
        #等待General可點擊
        click_element_xpath(gcadmin_general_xpath)
        
        #等待Game List可點擊
        if not click_element_xpath(game_list_xpath):
            #重新點General
            click_element_xpath(gcadmin_general_xpath)
            #再點Game List
            click_element_xpath(game_list_xpath)
            
        #跳出leftframe，並進入mainFrame
        switch_to_frame("mainFrame")
        
        #取得sheet內所有資料
        all_values = sheet.get_all_values()
        #資料從第二行開始
        data_rows = all_values[1:]
        
        #provider
        wait_chromeweb_id("ddlProviders")
        provider =  home_views.chromeWeb
        provider = provider.find_element(By.ID, "ddlProviders")
        
        #將資料轉換成字典，key = gametype
        ac_num = 1
        for row in data_rows:
            gameType = row[0]
            gameName = row[1]
            providerGameType = row[2]
            
            if gameType:
                home_views.chromeWeb.find_element(By.ID, "txtGameType").clear()
                home_views.chromeWeb.find_element(By.ID, "txtGameType").send_keys(gameType)
                if game == "pp":
                    Select(provider).select_by_value("22")  # 下拉選單取值 PP
                else:
                    Select(provider).select_by_value("6")  # 下拉選單取值 SG
                #點submit
                click_element_xpath(game_gameList_submit_xpath)
                time.sleep(2)
                gcadmin_gameType = home_views.chromeWeb.find_element(By.XPATH, '//*[@id="tbody1"]/tr/td[3]').text
                gcadmin_Name = home_views.chromeWeb.find_element(By.XPATH, '//*[@id="tbody1"]/tr/td[4]').text
                gcadmin_ProviderGameType = home_views.chromeWeb.find_element(By.XPATH, '//*[@id="tbody1"]/tr/td[5]').text
                if gcadmin_gameType == gameType and gcadmin_Name == gameName and gcadmin_ProviderGameType == providerGameType:
                    valueList.append(f'{ac_num}. {gameType} : {gameName} : {providerGameType} -- PASS')
                else:
                    valueList.append(f'{ac_num}. {gameType} : {gameName} : {providerGameType} -- FAIL')
                ac_num += 1
                
        #跳出mainFrame，並進入leftFrame
        switch_to_frame("leftFrame")
        return valueList
    
    else:
        pass
    
def excelMinBetLimit(sheet, cur_range, minRow):
    excelbetLimit = cur_range[minRow].value
    excelbetLimit = excelbetLimit.replace("=", "")
    try:
        minValue = sheet[excelbetLimit].value
    except:
        excelbetLimitList = excelbetLimit.split("*")
        if isinstance(excelbetLimitList[0], int):
            minValue = sheet[excelbetLimitList[0]].value * int(excelbetLimitList[1])
        else:
            try:
                thisMin = sheet[excelbetLimitList[0]].value
            except:
                return 0
            try:
                thisMin = thisMin.replace("=", "")
                minValue = sheet[thisMin].value * int(excelbetLimitList[1])
            except:
                minValue = thisMin * int(excelbetLimitList[1])
    min = "%.2f" % minValue
    return min


def excelMaxBetLimit(sheet, cur_range, maxRow):
    excelbetLimit = cur_range[maxRow].value
    excelbetLimit = excelbetLimit.replace("=", "")
    try:
        maxValue = sheet[excelbetLimit].value
    except:
        excelbetLimitList = excelbetLimit.split("*")
        if isinstance(excelbetLimitList[0], int):
            maxValue = sheet[excelbetLimitList[0]].value * int(excelbetLimitList[1])
        else:
            try:
                thisMax = sheet[excelbetLimitList[0]].value
            except:
                return 0
            try:
                thisMax = thisMax.replace("=", "")
                maxValue = sheet[thisMax].value * int(excelbetLimitList[1])
            except:
                maxValue = thisMax * int(excelbetLimitList[1])
    max = "%.2f" % maxValue
    return max
 
def adminPage():
    cur_dict = {}
    cur_dict.clear()
    try:
        #一次性找出所有行
        rows = home_views.chromeWeb.find_elements(By.XPATH, '//*[@id="tablelist"]/tbody/tr')
        #print(f"找到 {len(rows)} 行資料\n")
        if len(rows) == 0:
            time.sleep(2)  # 等待2秒後重試
            rows = home_views.chromeWeb.find_elements(By.XPATH, '//*[@id="tablelist"]/tbody/tr')
            #print(f"找到 {len(rows)} 行資料\n")
        for row in rows:
            #定位td
            td = row.find_elements(By.TAG_NAME, 'td')
            #找出第2欄和第3欄的值
            gameName = home_views.chromeWeb.execute_script("return arguments[0].textContent ", td[1]).strip()
            betLimit = home_views.chromeWeb.execute_script("return arguments[0].textContent ", td[2]).strip()

            #將資料加入字典
            cur_dict[gameName] = betLimit
    except Exception as e:
        print(f"錯誤訊息: {e}")
        return cur_dict
    #for num in range(1, 101):  # 遍歷第一頁1~100行，取Game Name 和 Bet Limit
    #    try:
    #        gameName1 = home_views.chromeWeb.find_element(
    #            By.XPATH, f'//*[@id="tablelist"]/tbody/tr[{num}]/td[2]'
    #        ).text
    #        betLimit1 = home_views.chromeWeb.find_element(
    #            By.XPATH, f'//*[@id="tablelist"]/tbody/tr[{num}]/td[3]'
    #        ).text
    #    except:
    #        break
    #    cur_dict[gameName1] = betLimit1  # 取出值加入admin字典
    return cur_dict