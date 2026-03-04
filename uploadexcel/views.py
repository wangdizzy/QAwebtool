from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, redirect
from django.contrib import messages
from django.conf import settings
from .forms import ExcelUploadForm
from django.http import JsonResponse
from .models import UploadedExcel
from openpyxl import load_workbook
from io import BytesIO
from oauth2client.service_account import ServiceAccountCredentials
import os
import gspread


def upload_excel(request):
    if request.method == "POST":
        # 獲取上傳的檔案
        uploaded_file = request.FILES.getlist("excelFile") 
        
 
        # 只上傳1筆資料且檔名以"PP_"開頭，則執行excel_gcadmin
        if len(uploaded_file) == 1:
            fileName = uploaded_file[0].name
            if fileName[:3] == "PP_":
                excel_gcadmin(uploaded_file[0])
            else:
                clear_google_sheet() 
                excel_process(uploaded_file)
        else:
            # 清除Google Sheet中的資料
            clear_google_sheet()  
            excel_process(uploaded_file)
    else:
        form = ExcelUploadForm() 
    return render(request, "upload_excel.html")


def clear_google_sheet():
    sheet = google_sheet()
    all_values = sheet.get_all_values()
    
    #總行數
    total_rows = len(all_values)
    
    #設定清除範圍
    if total_rows > 1:
        range_to_clear = f"A2:Z{total_rows}"
        sheet.batch_clear([range_to_clear])
    
    
def excel_list(file_path):
    gameName_list = []
    # excel = load_workbook(file_path)  # 讀取excel檔
    for x in file_path:
        excel = load_workbook(x)
        for sheetName in excel.sheetnames:  # 把excel 的sheet全部取出遍歷
            try:
                sheet = excel[sheetName]  # 對EXCEL切換sheet
            except:
                break
            excelGame = sheet[1]
            excelGameName = excelGame[0].value
            if excelGameName is not None:
                gameNameRow = sheet[1]
                gameNameValue = gameNameRow[0].value
                gameName_list.append(gameNameValue)
    return gameName_list


def excel_gcadmin(file_path):
    
    #傳入google sheet name為GameType的worksheet
    gameData = google_sheet("GameType")
    
    #讀取excel檔案取出
    excel = load_workbook(file_path)
    
    #創建空列表
    updates = []

    #取得excel檔案的sheet1
    sheet = excel["Sheet1"]
    
    #創建空字典，儲存excel內gametype、provider game type、game name 
    excel_data = {}
    '''
    透過.iter_rows()方法遍歷工作表每一行
    min_row=2表示從第2行開始
    values_only=True表示只返回儲存格的值，而不是儲存格物件
    '''
    #從第2行開始
    empty_row = 2
    for row in sheet.iter_rows(min_row=2, values_only=True):
        gameType = row[0]
        providerGameType = row[1]
        excelGameName = row[2]
        if excelGameName:
            updates.append({
                'range':f'A{empty_row}:C{empty_row}',
                'values':[[gameType, excelGameName, providerGameType]]
            })
            empty_row += 1
            
    #如果all_updates有資料，存入sheet
    if updates:
        gameData.batch_update(updates)



def excel_process(uploaded_file):
    gameData = google_sheet()

    # 取得第A欄
    gameData_values = gameData.col_values(1)

    # 起始欄位從第2行開始
    empty_row = 2

    for x in uploaded_file:
        all_updates = []
        file_data = x.read()
        excel = load_workbook(filename=BytesIO(file_data))  # 讀取excel
        for excelSheet in excel.sheetnames:
            sheet = excel[excelSheet]  # 對EXCEL切換sheet
            excelGame = sheet[1]
            for x in range(1, 4):
                # 判斷min和max，excel檔案第2行
                minMaxcolumn = sheet[2]
                minMaxcolumnValue = minMaxcolumn[x].value
                if minMaxcolumnValue[0:3] == "Min":
                    minRow = int(x)
                elif minMaxcolumnValue[0:3] == "Max":
                    maxRow = int(x)
                else:
                    pass
            for column in range(3, 30):  # 對EXCEL的1~13行遍歷
                cur_range = sheet[column]  # EXCEL該sheet的行數 column=行號
                excel_game_currency = cur_range[0].value
                #將遊戲名稱寫入gameData(PP)的Game欄位
                excel_game_name = excelGame[0].value
                
                if excel_game_currency is None:
                    break
                elif excel_game_currency !="IDR2" and excel_game_currency !="VND2":
                    try:
                        min = (
                            "%.2f" % cur_range[minRow].value
                        )  # 抓取excel B欄位
                    except:
                        min = excelMinBetLimit(sheet, cur_range, minRow)
                    try:
                        max = (
                            "%.2f" % cur_range[maxRow].value
                        )  # 抓取excel C欄位
                    except:
                        max = excelMaxBetLimit(sheet, cur_range, maxRow)
                    excel_game_minmax = f'{min} ~ {max}'
                elif excel_game_currency == "IDR2":
                    #將遊戲幣別寫入gameData(PP)的Currency欄位
                    excel_game_currency = 'IDR'
                    try:
                        # excel文檔最小值 取小數後2位，沒有補0
                        min = "%.2f" % cur_range[minRow].value
                    except:
                        min = excelMinBetLimit(sheet, cur_range, minRow)
                    try:
                        # excel文檔最大值 取小數後2位，沒有補0
                        max = "%.2f" % cur_range[maxRow].value
                    except:
                        max = excelMaxBetLimit(sheet, cur_range, maxRow)
                    excel_game_minmax = f"{min} ~ {max}"
                    #break
                elif excel_game_currency == "VND2":
                    #將遊戲幣別寫入gameData(PP)的Currency欄位
                    excel_game_currency = 'VND'
                    try:
                        min = "%.2f" % cur_range[minRow].value  # 抓取excel B欄位
                    except:
                        min = excelMinBetLimit(sheet, cur_range, minRow)
                    try:
                        max = "%.2f" % cur_range[maxRow].value  # 抓取excel C欄位
                    except:
                        max = excelMaxBetLimit(sheet, cur_range, maxRow)
                    excel_game_minmax = f"{min} ~ {max}"
                    #break
                else:
                    break
                #將資料存入陣列，後續一次更新到sheet
                all_updates.append({
                    'range':f'A{empty_row}:C{empty_row}',
                    'values':[[excel_game_name, excel_game_currency, excel_game_minmax]]
                })
                empty_row += 1
            #如果all_updates有資料，存入sheet
            if all_updates:
                gameData.batch_update(all_updates)
    return all_updates
        


def excelMinBetLimit(sheet, cur_range, minRow):
    excelbetLimit = cur_range[minRow].value
    excelbetLimit = excelbetLimit.replace("=", "")
    try:
        minValue = sheet[excelbetLimit].value
    except:
        excelbetLimitList = excelbetLimit.split("*")
        if isinstance(excelbetLimitList[0], int):
            minValue = sheet[excelbetLimitList[0]].value * int(excelbetLimitList[1])
        else:
            try:
                thisMin = sheet[excelbetLimitList[0]].value
            except:
                return 0
            try:
                thisMin = thisMin.replace("=", "")
                minValue = sheet[thisMin].value * int(excelbetLimitList[1])
            except:
                minValue = thisMin * int(excelbetLimitList[1])
    min = "%.2f" % minValue
    return min


def excelMaxBetLimit(sheet, cur_range, maxRow):
    excelbetLimit = cur_range[maxRow].value
    excelbetLimit = excelbetLimit.replace("=", "")
    try:
        maxValue = sheet[excelbetLimit].value
    except:
        excelbetLimitList = excelbetLimit.split("*")
        if isinstance(excelbetLimitList[0], int):
            maxValue = sheet[excelbetLimitList[0]].value * int(excelbetLimitList[1])
        else:
            try:
                thisMax = sheet[excelbetLimitList[0]].value
            except:
                return 0
            try:
                thisMax = thisMax.replace("=", "")
                maxValue = sheet[thisMax].value * int(excelbetLimitList[1])
            except:
                maxValue = thisMax * int(excelbetLimitList[1])
        max = "%.2f" % maxValue
    return max


def google_sheet(worksheet_name="gameData(PP)"):

    sheet_json = {
        "type": "service_account",
        "project_id": "inlaid-sentinel-407402",
        "private_key_id": "e8289caa403c29ec8edeb62490cdb696984dea98",
        "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQC0FOWiThPW8PVC\nHxJT6zFuXVE5ekvsIDiDzznJRorYkZFRxV6n6H2c6rBUZQf2dhgYR6HorFcMz+oy\n9ihlO2Fv7fXD6bfn7xje56JPyVu3RXsmAB1K7eZYFjWYeULYe3Xl5eLZX4QYFSvI\nBK/CX+pboojywnPcJlSBxqW1q9cz+crvpfTXq5mPEdp1sdotOxmX6pTi8SvB4eSd\n3A31oCtpsEvXFwK1ZUwEVlw7cLdgGqhKrEgn8wHVrYiVYSA7pM19opy2gJJyqZVX\n6wzeGA1uBGTXDAEMBLJ8AStuFAjoG9/3OMay0Ln3ba9lc7vOWsdBFz/TqdSaRGre\nUJuDuFvJAgMBAAECggEAEfQOY8TIdhKeRPoPi/XD/xaHGUWfyZn5wGxZvL8PayFy\nHTah5fgIA+ui6jsLVO83nj4P/oAmCpU06mE/rD4EDBJrgN1tdA5SirCJrk4rGmWv\nLh3vTa/Tmd8W901JlIcUfTfSyqya4QLFU2LiOezxkrKs2BT6U42vuaN6FFdeNGRt\nPLPxSnNHwpv59XrCfFoOp78oU3LrE+J7DQrjITBSy3iVZQypW+x5JapJ9R5Ix5Ox\njHehNaqTJ4K5y31ZBEur2ZQTNEd95i+gBIHAJbb3L49C2zt0TwFbEVkdOeG5zZ3l\nJr9oPRYQN5QwlgZgIW8T5SKlRA8rsv7wAwm9qV/XXQKBgQDoVrQiTafWuLD1VHxJ\n3uMsI+dzMFRBKGmd8tbFD5V1atFrnXSTVBtg2RdsAM21WIG8D0EhLD5Kc7XlpIPv\n9sMWEaKKbGLk8a4eKXAEetsJjnnCdCLb+GqeO2Qp0JDZ6cKwZFKNopcDEEiEX39i\nZtHL82Eq7DbnkLJ0G8dwqZJB7QKBgQDGa80FkScVnyw5G0PEqtQ1wAfytJyGB3V6\nJDNdaC8XXm2vzUHWZM/Tyrw+qeHZRuLDu2gb2p221zE7CyYkxgj66VWBNRPnpEiY\nc3VV7bXBfVEAiftVRWAHgrHtGCRACKKt4OgfrJrzRAcekye4Lm4er4QDEZI4uBZq\nYm8fqVG1zQKBgQDegDoeJ9Q2M8V0DKbCb6uK2A+NJplplPQgiTDYo2X0folzz+SW\nOxPFGeHuUo6tvsbvfIRY6m/1CP8HnxejNOP7PIQ2oDnNGw4uYGygPa+KZWGBsYq4\nshwY0LPJv60Yo18JYeoVLcIE8xEfg/0QFXuRH9DMNE8YUGA2BWxoHlysuQKBgQCl\nRMcpGuTeGo1gJ3iDx/InrwIvwwYYkP/ls19hLtUCdvGPm7x50dBVTSkMXL20F1nr\nxB4MDUSONaFY14l22chDDbTdgRNKPskEyi5yWyOnvTSJ6WQBe15oAxEmNZSEDW1K\nvOk68K7DbucrLVDJFUs9nd2sHKeZPKPXCpQaYBKiBQKBgEdAQVOL6p5sm/acvKsF\nSzzieDOzCeeiGYmeqguiEg7sLJ5UcSlCdKDGKFtgc1vi9j1cVcAah2PBgigyuLhx\nrkHfPNKePKZ/Do3WTuG2HOmlRXK/FJXsE+9m6joD9ZVipE2mTowWNq675EkUkfrd\n/Ef2Q0gOIPY2sY5/aKQKR18P\n-----END PRIVATE KEY-----\n",
        "client_email": "qa-451@inlaid-sentinel-407402.iam.gserviceaccount.com",
        "client_id": "115465534078375635707",
        "auth_uri": "https://accounts.google.com/o/oauth2/auth",
        "token_uri": "https://oauth2.googleapis.com/token",
        "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
        "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/qa-451%40inlaid-sentinel-407402.iam.gserviceaccount.com",
        "universe_domain": "googleapis.com",
    }

    # 定義存取的範圍 feeds = google sheet
    scopes = ["https://spreadsheets.google.com/feeds"]

    # 將金鑰放在dict使用
    credentials = ServiceAccountCredentials.from_json_keyfile_dict(sheet_json, scopes)

    # 傳入gspread模組
    client = gspread.authorize(credentials)

    # 使用open_by_key方式傳入google sheet 金鑰
    spreadsheet = client.open_by_key("1qWdc0QTGY13LEsr_N_5SA4cfyWTHaGXk8GhJFosoNzc")

    # 指定執行頁面
    gameData = spreadsheet.worksheet(worksheet_name)

    return gameData
